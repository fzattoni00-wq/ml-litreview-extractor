#!/usr/bin/env python3
"""
write_excel.py — Deterministic, NO-AI Excel writer for the battery ML literature review.

Takes the verified JSON rows (output of the extraction workflow) and writes them into a
COPY of the model template, preserving the original 28 columns and adding 6 analytical
columns. The original .xlsx is NEVER modified.

Legal-safety conventions:
  - Cells listed in a row's "uncertain_fields" are highlighted YELLOW and collected into
    the "Fields to verify (manual)" column, so a human knows exactly what to double-check.
  - Missing data must already be "Not reported" (the agents are instructed never to invent).

Usage:
  python write_excel.py <verified_rows.json> <template.xlsx> <output.xlsx>

verified_rows.json = list of records; each record is either a row dict (33 keys) OR a
verifier record {pdf_id, row:{...}, overall_confidence, uncertain_fields, ...}.
"""
import sys, json, copy
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

# (Excel header, json key) in final order. First 28 == original template columns.
COLUMNS = [
    ("PDF name", "pdf_name"),
    ("Paper title", "paper_title"),
    ("Authors", "authors"),
    ("Year", "year"),
    ("Publisher / venue", "publisher_venue"),
    ("DOI / link", "doi_link"),
    ("Target task", "target_task"),
    ("Target definition", "target_definition"),
    ("Battery chemistry", "battery_chemistry"),
    ("Cell / module / pack level", "cell_module_pack_level"),
    ("Dataset", "dataset"),
    ("Public dataset?", "public_dataset"),
    ("Experimental conditions", "experimental_conditions"),
    ("Input data type", "input_data_type"),
    ("Data preparation", "data_preparation"),
    ("Feature extraction", "feature_extraction"),
    ("Model category", "model_category"),
    ("NN architecture", "nn_architecture"),
    ("Training strategy", "training_strategy"),
    ("Validation strategy", "validation_strategy"),
    ("Baseline models", "baseline_models"),
    ("Metrics reported", "metrics_reported"),
    ("Main results", "main_results"),
    ("Generalisation tested?", "generalisation_tested"),
    ("Code/data available?", "code_data_available"),
    ("Advantages", "advantages"),
    ("Limitations", "limitations"),
    ("Relevance to my review", "relevance"),
    ("Notes", "notes"),
    # --- new analytical columns ---
    ("Architecture (plain-language)", "architecture_plain"),
    ("Input representation", "input_representation"),
    ("Results (structured)", "results_structured"),
    ("Results source + page", "results_source_page"),
    ("Extraction confidence", "extraction_confidence"),
    ("Fields to verify (manual)", "_fields_to_verify"),
]

YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
HDR_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HDR_FONT = Font(color="FFFFFF", bold=True)
KEY_TO_COLIDX = {k: i + 1 for i, (_, k) in enumerate(COLUMNS)}


def normalise(rec):
    """Return (row_dict, uncertain_list, overall_conf) from either flat row or verifier record."""
    if isinstance(rec, dict) and "row" in rec and isinstance(rec["row"], dict):
        row = dict(rec["row"])
        unc = rec.get("uncertain_fields") or row.get("uncertain_fields") or []
        conf = rec.get("overall_confidence") or row.get("extraction_confidence")
        if conf:
            row["extraction_confidence"] = conf
        return row, list(unc)
    row = dict(rec)
    return row, list(row.get("uncertain_fields") or [])


def main():
    if len(sys.argv) != 4:
        print(__doc__)
        sys.exit(1)
    rows_path, template_path, out_path = sys.argv[1:4]

    data = json.load(open(rows_path, encoding="utf-8"))
    if isinstance(data, dict) and "rows" in data:
        data = data["rows"]

    wb = openpyxl.load_workbook(template_path)
    if "Review_Matrix" in wb.sheetnames:
        del wb["Review_Matrix"]
    ws = wb.create_sheet("Review_Matrix", 0)

    # header
    for c, (title, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=c, value=title)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"

    # rows
    r = 2
    n_flagged_cells = 0
    for rec in data:
        row, uncertain = normalise(rec)
        uncertain_set = set(uncertain)
        for title, key in COLUMNS:
            c = KEY_TO_COLIDX[key]
            if key == "_fields_to_verify":
                val = "; ".join(uncertain) if uncertain else ""
            else:
                val = row.get(key, "")
                if isinstance(val, list):
                    val = "; ".join(str(x) for x in val)
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if key in uncertain_set:
                cell.fill = YELLOW
                n_flagged_cells += 1
        r += 1

    # sensible column widths, keyed by header title (robust to reordering)
    from openpyxl.utils import get_column_letter
    widths_by_title = {
        "PDF name": 34, "Paper title": 40, "Authors": 34, "Year": 8,
        "Publisher / venue": 24, "DOI / link": 30, "Feature extraction": 40,
        "Architecture (plain-language)": 50, "Results (structured)": 36,
        "Results source + page": 30, "Extraction confidence": 16,
        "Fields to verify (manual)": 30,
    }
    for idx, (title, _) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths_by_title.get(title, 26)

    # provenance note appended to Field_Guide if present
    if "Field_Guide" in wb.sheetnames:
        fg = wb["Field_Guide"]
        nr = fg.max_row + 2
        fg.cell(row=nr, column=1, value="v2 pipeline (Clara)")
        fg.cell(row=nr, column=2, value=(
            "Re-extracted from full text + tables + VISION on rendered result pages. "
            "6 new columns added (Architecture plain-language, Input representation, "
            "Results structured, Results source+page, Extraction confidence, Fields to verify). "
            "Yellow cells = flagged by the adversarial verifier for manual check before publication. "
            "Absent data = 'Not reported' (never invented)."))

    wb.save(out_path)
    print(f"Wrote {r-2} rows x {len(COLUMNS)} cols -> {out_path}")
    print(f"Yellow-flagged cells needing manual verification: {n_flagged_cells}")


if __name__ == "__main__":
    main()
